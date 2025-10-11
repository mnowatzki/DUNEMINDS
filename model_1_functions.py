## TO DO: fun_array_dataset: create list of the images that are added to array so they can be checked later

########## import packages & libraries ##########

import os
import sys
from osgeo import gdal
import random
import rasterio
import numpy as np
import csv
import pandas as pd
import geopandas as gpd
import shutil
import tensorflow as tf

import openpyxl
from openpyxl import load_workbook

import ee

from rasterio.warp import reproject
from rasterio.enums import Resampling


# to make sure screen on linux cluster saves figures etc even if window is closed
import matplotlib
matplotlib.use('Agg')


######### function to check if folder exists and contains a specific type of files ##########

def fun_check_folder_file(path, file_suffix):
    print(f"path: {path}")
    print(f"file suffix: {file_suffix}")
    

    stop_var_fun_check_folder_file = 0

    if os.path.exists(path):
    
        for files in os.listdir(path):
            if files.endswith(file_suffix):
                stop_var_fun_check_folder_file += 1
                       
    else:
        print(path)
        sys.exit("STOP. Path not existing.")
        
    return stop_var_fun_check_folder_file


########## function to check if file exists ##########

def fun_check_file(path, file_suffix, path_and_file):
    
    stop_var_fun_check_file = 0
   
    for files in os.listdir(path):
        if files.endswith(file_suffix):
            stop_var_fun_check_file += 1
            
    if stop_var_fun_check_file == 0:
        with open(path_and_file, 'w'):
            pass
        print("File created.")
    
    else:
        print("File already existing.")



########## check if folder exists and contains files with prefix ##########

def fun_check_folder_file_pref(path, file_suffix, file_prefix):
    

    stop_var_fun_check_folder_file = 0

    if os.path.exists(path):
    
        for files in os.listdir(path):
            if files.endswith(file_suffix) and files.startswith(file_prefix):
                stop_var_fun_check_folder_file += 1  
                       
    else:
        sys.exit("STOP. Path not existing.")
        
    return stop_var_fun_check_folder_file




########## function to check if a folder exists and create it if it does not ##########

def fun_check_create_folder(path):

    if os.path.exists(path):
        print("Folder exists.")
    
    else:
        os.mkdir(path)
        print("Created folder.")

    return path


########## function that asks user for input on whether to continue even though files have already been created

def fun_exit_decision():


    def fun_exit_decision_input():

        exit_decision_input = input("Files already existing. Would you like to proceed anyway? (y/n)")
        
        return exit_decision_input
    
    
    exit_decision = fun_exit_decision_input()
    
    while exit_decision != 'y' and exit_decision != 'n':
        print("Wrong input. Try again!")
        exit_decision = fun_exit_decision_input()
        
    if exit_decision == 'y':
        print("Let's keep going!")
                    
    elif exit_decision == 'n':
        sys.exit("Files already existing. You decided to stop.")




########## function to normalise tiles based on values of whole study area ##########
## pre-processing Sentinel-2 tiles and putting them into pp training folder as geotiff

def fun_normalisation(path_in,
                      normal,
                      min_max_creation,
                      path_download_images,
                      download_file_suffix,
                      csv_path_file,
                      file_suffix, 
                      dt,
                      path_out,
                      name_out,
                      dataprep_data,
                      study_area_dataprep):

    
    images = os.listdir(path_in)

    if normal: #only if normalisation is required

        ## Retrieval of the min max values of the whole study area
        if min_max_creation: #only if there are no min & max values for the entiry study area saved

            download_images = os.listdir(path_download_images)

            for i, image_name in enumerate(download_images): #go through the image tiles downloaded from GEE
                if image_name.endswith(download_file_suffix) == True: #only consider those with a particular suffix


                    with rasterio.open(path_download_images + '/' + image_name) as src: #open tiles
                        for j, band in enumerate(range(1, src.count + 1)): #go through each band of the currently open tile
                            df = pd.read_csv(csv_path_file, header=0, index_col=None) # read the csv file that holds the currently smalles min / largest max values
                            old_min = df.iloc[j,0] # current min value of that band of a previous tile saved in csv
                            #print(old_min)
                            old_max = df.iloc[j,1] # current max value of that band of a previous tile saved in csv
                            print(old_max)
                            band_data = src.read(band) # read band

                            min_value = np.nanmin(band_data) # calculate min value of the current tile / band
                            #print(min_value)
                            if min_value < old_min: # if current value is smaller than the saved old one - overwrite old
                                df.iloc[j,0] = min_value

                            max_value = np.nanmax(band_data) # calculate max value of the current tile / band
                            print(max_value)
                            if max_value > old_max: # if current value is larger than the saved old one - overwrite old
                                df.iloc[j,1] = max_value

                            df.to_csv(csv_path_file, index=False) # save new values to csv (overwrite)


        ## applying normalisation to training / vali / test data
        # read all images in folder holding original data
        for k, image_name in enumerate(images):
            if image_name.endswith(file_suffix) == True:
                print(image_name)
                img_pp = [] # empty list to store preprocessed array
                
                with rasterio.open(path_in + "/" + image_name) as src: # open original files
                    for m, band in enumerate(range(1, src.count + 1)): # go through bands individually

                        img_band = src.read(band) # read band as np array

                        # check if GTiff has been converted to ndarray
                        array_test = rasterio.dtypes.is_ndarray(img_band) 
                        if array_test == False:
                            sys.exit("STOP. GTiff has not been converted to ndarray!")

                        # convert to float if it isn't float already
                        if dt != 'no_conversion':
                            data_type = img_band.dtype
                            #print(data_type)

                        elif data_type != dt:
                            img_band = img_band.astype('f4')

                        df = pd.read_csv(csv_path_file, header=0, index_col=None) # read csv file with stored min max values of whole area
                        band_min = np.nanmin(img_band) # calculate min max values for current tile / band
                        band_max = np.nanmax(img_band)
                        overall_min = df.iloc[m,0] # get min max values of whole study area for currently open band  from csv file
                        overall_max = df.iloc[m,1]
                
                        if band_min < 0 or band_max > 1: # check if currently open tile / band is already normalised to 0-1 range
                            img_band_pp = (img_band - overall_min) / (overall_max - overall_min) # calculate NORMALISATION
                            #print(img_band_pp)
                            img_pp.append(np.array(img_band_pp)) # stack bands of each file to create multiband array

                        else:
                            img_pp.append(np.array(img_band)) # in case no normalisation is needed: just stack original arrays

                        

                    img_pp = np.array(img_pp) # create np array from img_pp list

                    ## calculate parameters for export
                    array_shape = img_pp.shape
                    #print(array_shape)
                    band_nr = array_shape[0]
                    #print("band nr: ", band_nr)
                    array_rows = array_shape[1]
                    #print("rows: ", array_rows)
                    array_columns = array_shape[2]
                    #print("columns: ", array_columns)


                    ## transformation into GeoTiff
                    with rasterio.open(path_out + "/" + 'pp_' + study_area_dataprep + '_' + image_name,
                                        'w',
                                        driver='Gtiff',
                                        width=array_rows,
                                        height=array_columns,
                                        count=band_nr,
                                        crs=src.crs,
                                        transform=src.transform,
                                        dtype='float32'
                                        ) as name_out:

                        name_out.write(img_pp)

    elif normal == False: # in case no normalisation is required (e.g. for binary labels) just copy original data into pp folder

        for file_name in images:
            if file_name.endswith(file_suffix):
                src_file = os.path.join(path_in, file_name)
                dst_file = os.path.join(path_out, file_name)

                shutil.copy(src_file, dst_file)




########## function to create array dataset of single images ##########

def fun_array_dataset(path_images_in):
    
    count = 0 # counter
    dataset_out = [] # create empty dataset
    
    images_in = os.listdir(path_images_in) # go to path containing images
    
    # read all images in folder as arrays and append
    for i, image_name in enumerate(sorted(images_in)):
        if image_name.endswith('label.tif'): 
            #print(image_name)
            image = rasterio.open(path_images_in + '/' + image_name)
            #show(image)
            image = image.read()
            image = image.transpose((1,2,0)) # order of dimensions is now size, size, bands
            #print(image)
            
            dataset_out.append(np.array(image))
      
            count += 1
            
            # counter
            #if count % 1 == 0:
                #print(count, " images added")
                
    #print(dataset_out)
    #print(images_in)
    print("SHAPE dataset:")
    print(dataset_out[1].shape)
    print("SIZE dataset:")
    print(len(dataset_out))
    
    
    # convert dataset (lists) to arrays
    #dataset_out = np.expand_dims(np.array(dataset_out), axis=1) # this is as done Sreeni's unet, but doesn't work here
    dataset_out = np.array(dataset_out)
    #print(dataset_out)
    print(dataset_out.shape) #number of tiffs, new axis, bands, size, size
    
    return dataset_out




########## function to read and print csv ##########

def fun_read_csv(file_path):

    with open (file_path, 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            print(row)



########## function to get matching images from two folders and create a training dataset ##########

def fun_concatenate_img(paths):

    # get images from each folder, sort them and only choose tif files
    data = [] #empty list that will be filled with a list of all the datsets used
    
    for path in paths:
        # create list of lists - every list item (one directory) contains list of files
        data.append(sorted([file for file in os.listdir(path) if file.endswith('.tif')]))

    dataset_out = [] #create empty dataset

    if len(set([len(item) for item in data])) > 1: #make sure that the same number of images is in both folders
        sys.exit('stop, different number of images per image type.')

    for i in range(len(data[0])): # goes through every item in data (each item is a list of files)
        
        img_list = [] #empty list that will be filled with rasterio.open commands for every  image
        
        for path in paths:

            img_path = path + '/' + data[paths.index(path)][i]

            # read and transpose image
            img = rasterio.open(img_path).read().transpose((1,2,0))

            # Exclude blue channel if input data is more than two items (e.g. optical and DEM) AND one of them is optical 4 channel
            num_channels = img.shape[2]
            if len(paths) > 1 and num_channels == 4: 
                img = img[:,:,[0,1,3]] #select R, G, NIR - excluding Blue channel


            img_list.append(img) # add to list that will later be used for concatenation


        # concatenate each single images - stack them on top of each other to create one more img channel
        conc_image = np.concatenate(img_list, axis = 2)

        # add concatenated image to training image stack
        dataset_out.append(np.array(conc_image))


    dataset_out = np.array(dataset_out)

    dataset_info = []  # List to store (path, file_name) tuples

    for path in paths:
        # ... (your existing code)

        for i in range(len(data[0])):
            # ... (your existing code)

            # Store the folder path and file name as a tuple
            dataset_info.append((path, data[paths.index(path)][i]))
    

    return dataset_out, dataset_info




########## function to save .py file as .txt file ##########

def save_py_txt(input_path: str, output_path: str):
    # Read the Python script
    with open(input_path, 'r') as file:
        script_content = file.read()

    # Save the script content into a text file
    with open(output_path, 'w') as file:
        file.write(script_content)



########## Function to create and add important input and result data to spreadsheet ##########

def workbook_creation_addition(path, variables):


    # Load the existing workbook or create a new one
    try:
        workbook = load_workbook(path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Select the active sheet
    sheet = workbook.active


    # Determine the next available row
    next_row = sheet.max_row + 1

    # Add header row if the sheet is empty
    if sheet.max_row == 1:
        header_row = ['RUN_ID', 
                    'DATE', 
                    'OPTIMIZER',
                    'LOSS',
                    'METRIC',
                    'ACTIVATION_1',
                    'ACTIVATION_2',
                    'DATA',
                    'STUDY_AREA',
                    'TRAIN_VALI_SPLIT',
                    'AUGMENT_DECISION',
                    'ROTATION_RANGE',
                    'WIDTH_SHIFT',
                    'HEIGHT_SHIFT',
                    'ZOOM_RANGE',
                    'HORIZONTAL_FLIP',
                    'VERTICAL_FLIP',
                    'BRIGHTNESS_RANGE',
                    'FILL_MODE',
                    'BRIGHTNESS_RANDOM_RANGE',
                    'TILE_SIZE_TRAIN',
                    'BATCH_SIZE',
                    'EPOCH_NR',
                    'SHUFFLE_TRAIN',
                    'EPOCH_MONITOR',
                    'EPOCH_PATIENCE',
                    'STEPS_PER_EPOCH',
                    'MONITOR_BESTWEIGHT',
                    'PRED_THRESHOLD',
                    'PRE_TRAIN_DECISION',
                    'PRE_TRAIN_RUN',
                    'EPOCHS_RUN',
                    'BEST_EPOCH_LOSS',
                    'BEST_EPOCH_VAL_LOSS',
                    'BEST_EPOCH_ACC',
                    'BEST_EPOCH_VAL_ACC',
                    'IOU',
                    'AVG_TRUE_POS',
                    'AVG_FALSE_POS',
                    'AVG_TRUE_NEG',
                    'AVG_FALSE_NEG'
                    ]
        sheet.append(header_row)


    # Append variables to the next row in the spreadsheet
    sheet.append(variables)

    # Save the modified spreadsheet
    workbook.save(path)




########## turning neighbour cells of 1 into 0.5 ##########

# def update_zeros_near_ones(arr):


#     instances = arr.shape[0]
#     rows = arr.shape[1]
#     cols = arr.shape[2]

#     if arr.dtype != np.float32:  # Check if the array is not already of float type
#         result = arr.astype(np.float32)  # Convert the array to float
#     else:
#         result = np.copy(arr)  # Create a copy of the float array

#     for i in range(instances):

#         for row in range(rows):

#             for col in range(cols):

#                 if arr[i, row, col] == 0:
#                     # Check 4-neighbourhood (left, right, top, bottom)
#                     if col > 0 and arr[i, row, col - 1] == 1:
#                         result[i, row, col] = 0.5
#                     elif col < cols - 1 and arr[i, row, col + 1] == 1:
#                         result[i, row, col] = 0.5
#                     elif row > 0 and arr[i, row - 1, col] == 1:
#                         result[i, row, col] = 0.5
#                     elif row < rows - 1 and arr[i, row + 1, col] == 1:
#                         result[i, row, col] = 0.5

#     return result


def update_zeros_near_ones(arr):

    instances = arr.shape[0]
    rows = arr.shape[1]
    cols = arr.shape[2]

    if arr.dtype != np.float32:  # Check if the array is not already of float type
        result = arr.astype(np.float32)  # Convert the array to float
    else:
        result = np.copy(arr)  # Create a copy of the float array

    for i in range(instances):

        for row in range(rows):

            for col in range(cols):

                if arr[i, row, col] == 0:
                    # Check 8-way neighborhood
                    for r in range(row - 1, row + 2):
                        for c in range(col - 1, col + 2):
                            if (0 <= r < rows) and (0 <= c < cols) and arr[i, r, c] == 1:
                                result[i, row, col] = 0.5

    return result





########## normalising images for plotting when their values are too close to each other to display as different values ##########

def plot_normalisation(image):

    image = np.array(image)

    unique_values = np.unique(image) #calculates every value occuring in the image

    a = 0.01
    b = 0.99

    if len(unique_values) == 1:
        # If the image contains only one unique value, set default values for min_val and max_val
        min_val = 0
        max_val = 1


    else:

        has_zero = np.any(unique_values == 0) #if zeros occur: has_zero
        has_one = np.any(unique_values == 1) #if ones occur: has_one

        
        if has_zero:
            min_val = unique_values[1] # zeros are turned into second lowest number

        else: 
            min_val = unique_values[0]
        

        if has_one:
            max_val = unique_values[-2] #ones are turned into second highest number

        else: 
            max_val = unique_values[-1]


    normalised_image = a + ((image - min_val) * (b - a)) / (max_val - min_val) #normalisation with min_val being stretched to 0.01 and max_val being stretched to 0.99

    normalised_image = np.clip(normalised_image, 0, 1)


    return normalised_image




########## function to get variables from txt file ##########

def get_list_var_from_txt(path, line_start):
    variable = None
    found_line = False

    with open(path, 'r') as file:
        for line in file:
            if line.startswith(line_start):
                variable = line[line.find("[")+1:line.find("]")] #extracts everything between the brackets
                print(variable)
                list_var = eval('[' + variable + ']') #adds brackets and creates list
                found_line = True
                break

    if not found_line:
        list_var = ['largest'] 

    return list_var


########## function to figure out what files are training and what are validation data after split ##########
def fun_identify_train_vali_files(path_train_csv, path_vali_csv, train_dataset, X_train, X_vali, train_image_info):

    # Find indices of images in X_train and X_vali within the concatenated dataset
    train_indices = [np.where((train_dataset == data).all(axis=(1, 2, 3)))[0][0] for data in X_train]
    vali_indices = [np.where((train_dataset == data).all(axis=(1, 2, 3)))[0][0] for data in X_vali]


    # Find corresponding image information for X_train and X_vali
    train_image_info_selected = [train_image_info[idx] for idx in train_indices]
    vali_image_info_selected = [train_image_info[idx] for idx in vali_indices]

    #print("Training image info:", train_image_info_selected)
    #print("Validation image info:", vali_image_info_selected)


    # Saving image information for X_train to a CSV file
    #with open('/soge-home/users/hert6209/model_1_data/csv/train_image_info.csv', 'w', newline='') as csvfile:
    with open(path_train_csv, 'w', newline='') as csvfile:
        csv_writer = csv.writer(csvfile)
        csv_writer.writerow(['Folder Path', 'File Name'])
        csv_writer.writerows(train_image_info_selected)

    # Saving image information for X_vali to a CSV file
    #with open('/soge-home/users/hert6209/model_1_data/csv/vali_image_info.csv', 'w', newline='') as csvfile:
    with open(path_vali_csv, 'w', newline='') as csvfile:
        csv_writer = csv.writer(csvfile)
        csv_writer.writerow(['Folder Path', 'File Name'])
        csv_writer.writerows(vali_image_info_selected)



########## function to calculate tiered IoU ##########

def calculate_tiered_iou(sec_tier_value: float, prediction_data, ground_truth_data):

    prediction_data_upd = update_zeros_near_ones(prediction_data)
    ground_truth_data_upd = update_zeros_near_ones(ground_truth_data)

    print(prediction_data_upd.shape)

    # Calculating full intersection
    intersection_full = ((prediction_data_upd == 1.0) & (ground_truth_data_upd == 1.0)) | ((prediction_data_upd == 0.5) & (ground_truth_data_upd == 0.5))

    # calculating limited intersection
    intersection_lim = (((prediction_data_upd == 1.0) & (ground_truth_data_upd == 0.5)) | ((prediction_data_upd == 0.5) & (ground_truth_data_upd == 1.0)))

    union_tiered = ((prediction_data_upd == 0.5) | (ground_truth_data_upd == 0.5) | (prediction_data_upd == 1.0) | (ground_truth_data_upd == 1.0))


    IoU_tiered = (np.sum(intersection_full) + ((np.sum(intersection_lim))*sec_tier_value)) / np.sum(union_tiered)

    print('IoU results for sec_tier_value: ', sec_tier_value)

    print('mean intersection_full: ', np.mean(intersection_full))
    print('sum intersection_full: ', np.sum(intersection_full))
    print('mean intersection_lim: ', np.mean(intersection_lim))
    print('sum intersection_lim: ', np.sum(intersection_lim))
    print('mean union_tiered: ', np.mean(union_tiered))
    print('sum union_tiered: ', np.sum(union_tiered))
    print('IoU_tiered NEW: ', IoU_tiered)

    return IoU_tiered
    
    
    
    
########## function to prepare very large image to be plotted including reprojection ##########



def prep_plot_large_image(downscale_factor, path):
    
    with rasterio.open(path) as src:
        
        data = dataset.read(
            out_shape=(
                src.count,
                int(src.height * downscale_factor),
                int(src.width * downscale_factor)
            ),
            resampling=Resampling.bilinear
                
        )
        
        # scale image transform
        transform = src.transform * src.transform.scale(
            (src.width / data.shape[-1]),
            (src.height / data.shape[-2])
        )


    # Extract color channels
    img_blue = data[0, :, :]
    img_green = img_orig[1, :, :]
    img_red = img_orig[2, :, :]
    
    # Stack the channels
    img = np.stack([img_red, img_green, img_blue], axis=-1)
    
    # Print the shape of the resampled image
    print(img.shape)
    
    return img




